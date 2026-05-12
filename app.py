# app.py  -  Cotizador NC  |  Automatzia Brand  |  Mobile-first
# Deps: pip install streamlit pandas openpyxl xlrd

import base64
import os
import re
import smtplib
import ssl
from datetime import date
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_PRICE_FILE = "www/LISTA_DE_PRECIOS_2025.xlsx"
BRAND_DIR = Path("automatzia_brandkit")

# ─────────────────────────────────────────────────────────────────────────────
# BRAND ASSETS
# ─────────────────────────────────────────────────────────────────────────────
def _svg_b64(path):
    data = path.read_bytes()
    return base64.b64encode(data).decode()

LOGO_B64 = _svg_b64(BRAND_DIR / "logo_dark.svg") if (BRAND_DIR / "logo_dark.svg").exists() else ""
ICON_B64 = _svg_b64(BRAND_DIR / "icon.svg") if (BRAND_DIR / "icon.svg").exists() else ""

BRAND_CSS = """
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Poppins:wght@600;700&display=swap');
  :root {
    --navy:  #0B1F3A;
    --blue:  #00AEEF;
    --cyan:  #00FFC6;
    --gray1: #F0F4F8;
    --gray2: #CBD5E1;
    --text:  #1E293B;
    --card:  #FFFFFF;
    --radius: 14px;
    --shadow: 0 2px 16px 0 rgba(11,31,58,.10);
  }
  html, body, [class*="css"] {
    font-family: 'Inter', sans-serif !important;
    color: var(--text) !important;
    background: var(--gray1) !important;
  }
  #MainMenu, footer, header { visibility: hidden; }
  .block-container {
    padding: 0 !important;
    max-width: 480px !important;
    margin: 0 auto !important;
  }
  .nc-topbar {
    background: var(--navy);
    padding: 14px 20px 12px 20px;
    display: flex;
    align-items: center;
    gap: 12px;
    position: sticky;
    top: 0;
    z-index: 100;
    box-shadow: 0 2px 12px rgba(0,0,0,.25);
  }
  .nc-topbar img { height: 38px; }
  .nc-topbar-title {
    font-family: 'Poppins', sans-serif !important;
    font-size: 17px !important;
    font-weight: 700 !important;
    color: #F5F7FA !important;
    line-height: 1.1 !important;
  }
  .nc-topbar-sub {
    font-size: 10px !important;
    color: var(--cyan) !important;
    font-weight: 500 !important;
    letter-spacing: .04em !important;
  }
  .nc-card {
    background: var(--card);
    border-radius: var(--radius);
    box-shadow: var(--shadow);
    padding: 18px 16px 14px 16px;
    margin: 12px 12px 0 12px;
  }
  .nc-card-title {
    font-family: 'Poppins', sans-serif !important;
    font-size: 13px !important;
    font-weight: 700 !important;
    text-transform: uppercase;
    letter-spacing: .08em;
    color: var(--navy) !important;
    margin-bottom: 10px !important;
    display: flex;
    align-items: center;
    gap: 6px;
  }
  .nc-card-title::after {
    content: "";
    flex: 1;
    height: 2px;
    background: linear-gradient(90deg, var(--blue), var(--cyan));
    border-radius: 2px;
  }
  .nc-badge-ok   { background:#D1FAE5; color:#065F46; border-radius:8px; padding:8px 12px; font-size:13px; font-weight:600; }
  .nc-badge-warn { background:#FEF3C7; color:#92400E; border-radius:8px; padding:8px 12px; font-size:13px; font-weight:600; }
  div[data-testid="stButton"] > button[kind="primary"] {
    background: linear-gradient(135deg, var(--blue) 0%, var(--cyan) 100%) !important;
    color: var(--navy) !important;
    font-weight: 700 !important;
    font-size: 15px !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 12px 0 !important;
    box-shadow: 0 4px 14px rgba(0,174,239,.30) !important;
  }
  div[data-testid="stDownloadButton"] > button {
    background: var(--navy) !important;
    color: var(--cyan) !important;
    font-weight: 700 !important;
    font-size: 15px !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 12px 0 !important;
    width: 100% !important;
  }
  div[data-baseweb="input"] input,
  div[data-baseweb="textarea"] textarea {
    border-radius: 10px !important;
    border: 1.5px solid var(--gray2) !important;
    font-size: 14px !important;
    background: #FAFBFC !important;
  }
  div[data-baseweb="input"] input:focus,
  div[data-baseweb="textarea"] textarea:focus {
    border-color: var(--blue) !important;
    box-shadow: 0 0 0 3px rgba(0,174,239,.15) !important;
  }
  .stDataFrame { border-radius: 10px; overflow: hidden; }
  details summary { font-weight: 600 !important; font-size: 14px !important; color: var(--navy) !important; }
  .nc-spacer { height: 80px; }
  hr { border: none; border-top: 2px solid var(--gray1); margin: 12px 0; }
</style>
"""

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def normalize_model(x):
    return re.sub(r"\s+", " ", str(x).strip()).upper()


def find_header_row(file_path, sheet_name):
    raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
    for i, row in raw.iterrows():
        values = [str(v).strip().upper() for v in row if pd.notna(v)]
        if "MODELO" in values or "MODEL" in values or "CLAVE" in values or "SKU" in values:
            return i
    return 0


def load_price_df(file_path, sheet_name):
    hdr = find_header_row(file_path, sheet_name)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=hdr)
    df = df.loc[:, ~df.columns.str.match(r"^Unnamed")]
    df.dropna(how="all", inplace=True)
    return df


def parse_order_text(text):
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    rows = []
    for ln in lines:
        m = re.match(r"^(.*?)-(\d+)$", ln)
        if m:
            modelo = normalize_model(m.group(1))
            cajas = int(m.group(2))
            if cajas > 0:
                rows.append({"modelo": modelo, "cajas": cajas})
    if not rows:
        return pd.DataFrame(columns=["modelo", "cajas"])
    df = pd.DataFrame(rows)
    return df.groupby("modelo", as_index=False)["cajas"].sum()


def to_number(val):
    try:
        s = re.sub(r"[^\d.\-]", "", str(val))
        return float(s) if s else None
    except Exception:
        return None


def build_quote(order_df, price_df):
    empty = {"data": pd.DataFrame(), "missing": pd.DataFrame()}
    if order_df.empty:
        return {"ok": False, "msg": "Sin lineas validas. Formato: MODELO-CANTIDAD", **empty}

    nms = {c.strip().lower(): c for c in price_df.columns}

    def pick(candidates):
        for c in candidates:
            if c in nms:
                return nms[c]
        return None

    col_model   = pick(["modelo", "model", "clave", "codigo", "codigo", "sku"])
    col_desc    = pick(["articulo", "articulo", "descripcion", "descripcion", "description", "producto"])
    col_pzxcaja = pick(["cant x caja", "cant_x_caja", "cantxcaja", "pzs_x_caja", "pzxcaja",
                        "pzsxcaja", "piezasxcaja", "piezas_x_caja", "pzs caja", "pzs/caja"])
    col_price3  = pick(["precio 3", "precio3", "precio_3", "p3", "price3", "price 3"])

    if any(v is None for v in [col_model, col_desc, col_pzxcaja, col_price3]):
        return {"ok": False, "msg": "No se detectaron columnas necesarias en la lista de precios.", **empty}

    pw = price_df[[col_model, col_desc, col_pzxcaja, col_price3]].copy()
    pw.columns = ["_modelo", "_desc", "_pzs", "_precio"]
    pw["_modelo"] = pw["_modelo"].apply(normalize_model)
    pw["_pzs"]    = pw["_pzs"].apply(to_number)
    pw["_precio"] = pw["_precio"].apply(to_number)

    merged = order_df.merge(pw, left_on="modelo", right_on="_modelo", how="left")
    merged["total_pzs"] = merged["cajas"] * merged["_pzs"]
    merged["importe"]   = merged["total_pzs"] * merged["_precio"]

    bad     = merged["_desc"].isna() | merged["_pzs"].isna() | merged["_precio"].isna()
    missing = merged[bad][["modelo", "cajas"]].copy()
    found   = merged[~bad].copy()

    out = found.rename(columns={
        "modelo": "MODELO", "cajas": "CAJAS", "_desc": "DESCRIPCION",
        "_pzs": "PZS/CAJA", "total_pzs": "TOTAL PZS",
        "_precio": "PRECIO", "importe": "IMPORTE",
    })[["MODELO", "CAJAS", "DESCRIPCION", "PZS/CAJA", "TOTAL PZS", "PRECIO", "IMPORTE"]]

    if not missing.empty:
        return {"ok": False,
                "msg": "No cotizados: " + ", ".join(missing["modelo"].tolist()),
                "data": out, "missing": missing}
    return {"ok": True, "msg": "OK", "data": out, "missing": pd.DataFrame()}


def write_quote_xlsx(header, quote_df, missing_df=None, include_iva=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotizacion"

    navy_fill  = PatternFill("solid", fgColor="0B1F3A")
    blue_fill  = PatternFill("solid", fgColor="00AEEF")
    alt_fill   = PatternFill("solid", fgColor="EFF6FF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="CBD5E1")
    b_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    ws["A1"] = "COTIZACION"
    ws["A1"].font      = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    ws["A1"].fill      = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    fields = [
        ("FECHA",      header.get("fecha",      "")),
        ("CLIENTE",    header.get("cliente",    "")),
        ("DIRECCION",  header.get("direccion",  "")),
        ("VENDEDOR",   header.get("vendedor",   "")),
        ("TRANSPORTE", header.get("transporte", "")),
    ]
    for i, (lbl, val) in enumerate(fields):
        r = 2 + i
        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font      = Font(bold=True, name="Calibri", color="0B1F3A")
        lc.fill      = alt_fill
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=r, column=2, value=val)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=7)
        ws.row_dimensions[r].height = 17

    start_tbl = 2 + len(fields) + 2
    col_names  = ["MODELO", "CAJAS", "DESCRIPCION", "PZS/CAJA", "TOTAL PZS", "PRECIO", "IMPORTE"]
    col_widths = [14, 8, 40, 10, 11, 13, 13]
    for ci, (h, w) in enumerate(zip(col_names, col_widths), 1):
        c = ws.cell(row=start_tbl, column=ci, value=h)
        c.font      = Font(bold=True, color="0B1F3A", size=10, name="Calibri")
        c.fill      = blue_fill
        c.border    = b_all
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[start_tbl].height = 20

    for ri, row_vals in enumerate(quote_df.itertuples(index=False), 1):
        r    = start_tbl + ri
        fill = white_fill if ri % 2 else alt_fill
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border    = b_all
            c.fill      = fill
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if ci in (2, 4, 5, 6, 7) else "left")
            if ci in (6, 7):
                c.number_format = '"$"#,##0.00'
            elif ci in (2, 4, 5):
                c.number_format = "0"
        ws.row_dimensions[r].height = 16

    subtotal = float(quote_df["IMPORTE"].sum())
    iva      = subtotal * 0.16 if include_iva else 0.0
    total    = subtotal + iva
    n        = len(quote_df)
    r_tot    = start_tbl + n + 2

    totals_rows = [
        (0, "SUBTOTAL", subtotal, "0B1F3A"),
    ]
    if include_iva:
        totals_rows.append((1, "IVA 16%", iva, "0B1F3A"))
    totals_rows.append((len(totals_rows), "TOTAL", total, "00AEEF"))

    for offset, lbl, val, clr in totals_rows:
        lc = ws.cell(row=r_tot + offset, column=5, value=lbl)
        lc.font      = Font(bold=True, color=clr, name="Calibri")
        lc.alignment = Alignment(horizontal="right")
        vc = ws.cell(row=r_tot + offset, column=7, value=val)
        vc.font          = Font(bold=True, color=clr, name="Calibri")
        vc.number_format = '"$"#,##0.00'

    if missing_df is not None and not missing_df.empty:
        r_miss = r_tot + 4
        c = ws.cell(row=r_miss, column=1, value="NO COTIZADOS")
        c.font = Font(bold=True, color="B45309")
        for ci, h in enumerate(["MODELO", "CAJAS"], 1):
            ws.cell(row=r_miss + 1, column=ci, value=h).font = Font(bold=True)
        for ri2, rv in enumerate(missing_df.itertuples(index=False), 2):
            for ci, v in enumerate(rv, 1):
                ws.cell(row=r_miss + ri2, column=ci, value=v).border = b_all

    ws.freeze_panes = ws.cell(row=start_tbl + 1, column=1)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def send_email(to_addr, subject, body, attachment_bytes, attachment_name,
               smtp_host, smtp_port, smtp_user, smtp_pass, from_name=""):
    try:
        msg = MIMEMultipart()
        sender = f"{from_name} <{smtp_user}>" if from_name else smtp_user
        msg["From"]    = sender
        msg["To"]      = to_addr
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(attachment_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{attachment_name}"')
        msg.attach(part)
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ctx) as server:
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, to_addr, msg.as_string())
        return True, "Correo enviado correctamente."
    except Exception as e:
        return False, str(e)


# ─────────────────────────────────────────────────────────────────────────────
# PAGE SETUP
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Cotizador NC",
    page_icon="data:image/svg+xml;base64," + ICON_B64 if ICON_B64 else "📋",
    layout="centered",
    initial_sidebar_state="collapsed",
)
st.markdown(BRAND_CSS, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
for _k, _v in [
    ("quote_result", None),
    ("price_df", None),
    ("price_source", DEFAULT_PRICE_FILE),
    ("price_sheet", None),
    ("xlsx_bytes", None),
    ("include_iva", True),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ─────────────────────────────────────────────────────────────────────────────
# TOP BAR
# ─────────────────────────────────────────────────────────────────────────────
logo_img = (
    f'<img src="data:image/svg+xml;base64,{LOGO_B64}" />'
    if LOGO_B64
    else '<span style="font-size:22px;font-weight:700;color:white;">Automatzia</span>'
)
st.markdown(
    f"""
    <div class="nc-topbar">
      {logo_img}
      <div>
        <div class="nc-topbar-title">Cotizador</div>
        <div class="nc-topbar-sub">Automatiza &middot; Optimiza &middot; Transforma</div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# LOAD PRICE LIST
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def _load(path, sheet):
    return load_price_df(path, sheet)


def _refresh_price(path):
    try:
        sheets = pd.ExcelFile(path).sheet_names
        sheet  = sheets[0]
        df     = _load(path, sheet)
        st.session_state.price_df     = df
        st.session_state.price_sheet  = sheet
        st.session_state.price_source = path
        return True, sheet, df
    except Exception as e:
        return False, None, str(e)


if st.session_state.price_df is None and os.path.exists(DEFAULT_PRICE_FILE):
    _refresh_price(DEFAULT_PRICE_FILE)

# ─────────────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────────────
tab_cotizar, tab_precios, tab_ajustes = st.tabs(["📋 Cotizar", "🗂️ Lista de Precios", "⚙️ Ajustes"])

# ═══════════════════════════════════════════════════════════════
# TAB 1 – COTIZAR
# ═══════════════════════════════════════════════════════════════
with tab_cotizar:
    st.markdown('<div class="nc-card"><div class="nc-card-title">📄 Datos del cliente</div>', unsafe_allow_html=True)
    cliente    = st.text_input("Cliente",    placeholder="Nombre del cliente")
    direccion  = st.text_input("Direccion",  placeholder="Direccion de entrega")
    col_v, col_t = st.columns(2)
    with col_v:
        vendedor   = st.text_input("Vendedor",   placeholder="Vendedor")
    with col_t:
        transporte = st.text_input("Transporte", placeholder="Transporte")
    fecha = st.date_input("Fecha", value=date.today())
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="nc-card" style="margin-top:12px;"><div class="nc-card-title">✏️ Pedido</div>', unsafe_allow_html=True)
    st.caption("Una linea por producto - Formato: **MODELO-CANTIDAD** (ej. `MC-48-2`)")
    order_text = st.text_area(
        "Pedido", value="", height=160,
        placeholder="MC-48-2\nMC-26-1\nMC-15-3",
        label_visibility="collapsed",
    )
    allow_partial = st.checkbox("Descargar aunque falten modelos", value=True)
    include_iva = st.checkbox("Incluir IVA (16%)", value=st.session_state.include_iva)
    if include_iva != st.session_state.include_iva:
        st.session_state.include_iva = include_iva
        st.session_state.xlsx_bytes  = None
    run = st.button("Generar cotizacion", type="primary", use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

    if run:
        if st.session_state.price_df is None:
            st.session_state.quote_result = {
                "ok": False, "msg": "Lista de precios no disponible.",
                "data": pd.DataFrame(), "missing": pd.DataFrame(),
            }
        else:
            order_df = parse_order_text(order_text)
            st.session_state.quote_result = build_quote(order_df, st.session_state.price_df)
            st.session_state.xlsx_bytes   = None

    res = st.session_state.quote_result

    if res is not None:
        if res["ok"]:
            st.markdown(
                f'<div class="nc-card" style="margin-top:12px;"><div class="nc-badge-ok">Cotizadas {len(res["data"])} lineas correctamente</div></div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'<div class="nc-card" style="margin-top:12px;"><div class="nc-badge-warn">⚠️ {res["msg"]}</div></div>',
                unsafe_allow_html=True,
            )

    if res is not None and not res["data"].empty:
        st.markdown('<div class="nc-card" style="margin-top:12px;"><div class="nc-card-title">👁️ Vista previa</div>', unsafe_allow_html=True)
        st.dataframe(
            res["data"].style.format({"PRECIO": "${:,.2f}", "IMPORTE": "${:,.2f}"}),
            use_container_width=True, hide_index=True,
        )
        subtotal = res["data"]["IMPORTE"].sum()
        iva      = subtotal * 0.16
        total    = subtotal + iva
        if st.session_state.include_iva:
            c1, c2, c3 = st.columns(3)
            c1.metric("Subtotal", f"${subtotal:,.2f}")
            c2.metric("IVA 16%",  f"${iva:,.2f}")
            c3.metric("Total",    f"${total:,.2f}")
        else:
            c1, c2 = st.columns(2)
            c1.metric("Subtotal", f"${subtotal:,.2f}")
            c2.metric("Total",    f"${subtotal:,.2f}")
        st.markdown("</div>", unsafe_allow_html=True)

    if res is not None and not res["data"].empty and (res["ok"] or allow_partial):
        if st.session_state.xlsx_bytes is None:
            header_data = {
                "fecha":      fecha.strftime("%d de %B del %Y"),
                "cliente":    cliente,
                "direccion":  direccion,
                "vendedor":   vendedor,
                "transporte": transporte,
            }
            st.session_state.xlsx_bytes = write_quote_xlsx(
                header_data, res["data"],
                missing_df=res["missing"] if not res["ok"] else None,
                include_iva=st.session_state.include_iva,
            )

        fname = f"cotizacion_{date.today().strftime('%Y%m%d')}_{(cliente or 'cliente').replace(' ','_')}.xlsx"

        st.markdown('<div class="nc-card" style="margin-top:12px;"><div class="nc-card-title">💾 Exportar</div>', unsafe_allow_html=True)
        st.download_button(
            label="⬇️  Descargar Excel",
            data=st.session_state.xlsx_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        with st.expander("📧 Enviar por correo"):
            # ── Leer config desde secrets (o fallback vacío) ──────────────
            _sec   = st.secrets.get("email", {})
            _shost = _sec.get("smtp_host", "smtp.gmail.com")
            _sport = int(_sec.get("smtp_port", 465))
            _suser = _sec.get("smtp_user", "")
            _spass = _sec.get("smtp_pass", "")
            _fname_cfg = _sec.get("from_name", "")
            _cfg_ok = bool(_suser and _spass)

            if _cfg_ok:
                st.markdown(
                    f'<div style="font-size:12px;color:#065F46;background:#D1FAE5;'
                    f'border-radius:8px;padding:6px 10px;margin-bottom:8px;">'
                    f'✅ Enviando desde <b>{_suser}</b></div>',
                    unsafe_allow_html=True,
                )
            else:
                st.warning("Configura el correo en `.streamlit/secrets.toml` para activar esta funcion.")

            to_addr = st.text_input("Destinatario", placeholder="cliente@email.com", key="email_to")
            email_subject = st.text_input(
                "Asunto",
                value=f"Cotizacion {cliente or ''} - {date.today().strftime('%d/%m/%Y')}",
                key="email_subj",
            )
            email_body = st.text_area(
                "Mensaje",
                value=f"Hola,\n\nAdjunto encontraras la cotizacion solicitada.\n\nSaludos,\n{vendedor or ''}",
                height=100, key="email_body",
            )

            send_btn = st.button("📤 Enviar correo", use_container_width=True, disabled=not _cfg_ok)
            if send_btn:
                if not to_addr:
                    st.warning("Ingresa el correo del destinatario.")
                else:
                    with st.spinner("Enviando…"):
                        ok_m, msg_m = send_email(
                            to_addr, email_subject, email_body,
                            st.session_state.xlsx_bytes, fname,
                            _shost, _sport, _suser, _spass, _fname_cfg,
                        )
                    if ok_m:
                        st.success("✅ " + msg_m)
                    else:
                        st.error("❌ " + msg_m)

        st.markdown("</div>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# TAB 2 – LISTA DE PRECIOS
# ═══════════════════════════════════════════════════════════════
with tab_precios:
    st.markdown('<div class="nc-card"><div class="nc-card-title">🗂️ Lista de Precios 2025</div>', unsafe_allow_html=True)
    price_source_label = os.path.basename(st.session_state.price_source or DEFAULT_PRICE_FILE)
    st.caption(f"Archivo activo: **{price_source_label}**")

    uploaded = st.file_uploader(
        "Reemplazar lista de precios (.xlsx)", type=["xlsx"], key="price_upload",
        help="Columnas requeridas: MODELO, ARTICULO, CANT X CAJA, PRECIO 3",
    )
    if uploaded is not None:
        with open(DEFAULT_PRICE_FILE, "wb") as f:
            f.write(uploaded.read())
        ok_u, _, result_u = _refresh_price(DEFAULT_PRICE_FILE)
        if ok_u:
            _load.clear()
            st.success(f"✅ Lista actualizada permanentemente: {uploaded.name} ({len(result_u)} filas)")
        else:
            st.error(f"❌ No se pudo leer el archivo: {result_u}")
    st.markdown("</div>", unsafe_allow_html=True)

    if st.session_state.price_df is not None:
        st.markdown('<div class="nc-card" style="margin-top:12px;"><div class="nc-card-title">✏️ Editar precios</div>', unsafe_allow_html=True)
        st.caption("Edita directamente. Los cambios se aplican a la sesion actual.")
        # Forzar tipos editables: object -> string, numerics mantienen su tipo
        _editable_df = st.session_state.price_df.copy()
        for _c in _editable_df.columns:
            if _editable_df[_c].dtype == object:
                _editable_df[_c] = _editable_df[_c].where(
                    _editable_df[_c].isna(), _editable_df[_c].astype(str)
                ).astype(pd.StringDtype())
        edited_df = st.data_editor(
            _editable_df,
            use_container_width=True, num_rows="dynamic", key="price_editor", hide_index=True,
        )
        if st.button("💾 Aplicar cambios a la sesion", use_container_width=True):
            st.session_state.price_df     = edited_df
            st.session_state.quote_result = None
            st.session_state.xlsx_bytes   = None
            st.success("✅ Cambios aplicados. Ve a Cotizar para regenerar.")
        buf_price = BytesIO()
        edited_df.to_excel(buf_price, index=False)
        buf_price.seek(0)
        st.download_button(
            "⬇️ Descargar lista actual (.xlsx)", data=buf_price.getvalue(),
            file_name="LISTA_DE_PRECIOS_editada.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# TAB 3 – AJUSTES
# ═══════════════════════════════════════════════════════════════
with tab_ajustes:
    st.markdown('<div class="nc-card"><div class="nc-card-title">⚙️ Sistema</div>', unsafe_allow_html=True)
    rows_loaded = len(st.session_state.price_df) if st.session_state.price_df is not None else 0
    st.markdown(f"""
- **Fecha:** {date.today().strftime('%d/%m/%Y')}
- **Archivo activo:** `{st.session_state.price_source or '-'}`
- **Filas cargadas:** {rows_loaded}
- **Hoja:** `{st.session_state.price_sheet or '-'}`
""")
    if st.button("🔄 Recargar lista por defecto", use_container_width=True):
        _load.clear()
        ok_r, _, _ = _refresh_price(DEFAULT_PRICE_FILE)
        if ok_r:
            st.success("✅ Lista recargada.")
            st.session_state.quote_result = None
            st.session_state.xlsx_bytes   = None
        else:
            st.error("❌ Archivo no encontrado.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="nc-card" style="margin-top:12px;"><div class="nc-card-title">ℹ️ Instrucciones</div>', unsafe_allow_html=True)
    st.markdown("""
**Cotizar**
1. Llena los datos del cliente.
2. Escribe el pedido: `MODELO-CANTIDAD` por linea. Ej: `MC-48-2`
3. Presiona **Generar cotizacion**.
4. Descarga el Excel o envialo por correo.

**Lista de Precios**
- Edita precios directamente en la tabla.
- Sube un `.xlsx` nuevo para reemplazar la lista activa.
- Descarga la lista editada.

**Correo**
- Gmail: usa *App Password* (no tu contrasena normal).
- Puerto 465 = SSL · Puerto 587 = TLS.
""")
    st.markdown("</div>", unsafe_allow_html=True)

st.markdown('<div class="nc-spacer"></div>', unsafe_allow_html=True)
